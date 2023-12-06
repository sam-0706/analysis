from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, send_from_directory
import os
import random
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

it_companies = [
    "JP Morgan Chase", "Celigo", "Micron", "Pega Systems", "Oracle", "Darwinbox",
    "Providence India", "Deloitte", "Barclays PPO", "Accolite Digital", "Deliveroo",
    "Accelerixe360", "Eappsys Info Solutions Pvt Ltd", "Bosch", "Goldman Sachs",
    "Incture", "Qualcomm India P Ltd", "Data Insights", "Microsoft PPO", "Thomson Reuters",
    "UST", "HSBC", "Keyloop"
]
non_it_companies = ["UTS", "Technicfmc", "Ctrls", "Accenture", "Kpmg", "Accolite", "Asian paints", "Rinex", "Byteridge", "Axis", "Nxtwave", "Teachnook", "Roamonix"]
@app.route('/', methods=['GET', 'POST'])
def index():
    return render_template('index.html')


@app.route('/register', methods=['POST'])
def register():
    # Extract registration details
    details = {
        'roll_number': request.form['roll_number'],
        'name': request.form['name'],
        'branch': request.form['branch'],
        'phone_number': request.form['phone_number'],
        'email_id': request.form['email_id'],
        'official_email_id': request.form['official_email_id']
    }

    # Check if the user with the same roll number already exists in user details
    if is_roll_number_registered(details['roll_number'], 'user_details.xlsx'):
        return "This roll number has already been registered click on left arrow to get back to login page", 400

    # Check if the user with the same roll number already exists in user credentials
    if is_roll_number_registered(details['roll_number'], 'user_credentials.xlsx'):
        return "This roll number has already been registered click on left arrow to get back to login page", 400

    # Generate unique password
    password = details['branch'] + str(random.randint(1000, 9999))

    # Save details and credentials
    save_user_details(details, 'user_details.xlsx')
    save_credentials(details['roll_number'], password, 'user_credentials.xlsx')

    # Show password to user and redirect to login
    return render_template('show_password.html', username=details['roll_number'], password=password)


@app.route('/dev.html')
def dev_info():
    return render_template('dev.html')


@app.route('/admin_portal.html', methods=['GET', 'POST'])
def admin_portal():
    # Read user data from the 'user_details.xlsx' file
    user_data_list = fetch_user_data('user_details.xlsx')

    # Read user credentials from the 'user_credentials.xlsx' file
    user_credentials_list = fetch_user_credentials('user_credentials.xlsx')

    # Get the selected branch from the form submission
    selected_branch = request.form.get('branch_filter')

    # Check if "All Branches" is selected or no branch is selected
    if not selected_branch or selected_branch == "All Branches":
        # Show all data when "All Branches" is selected or no branch is selected
        filtered_user_data_list = user_data_list
    else:
        # Filter user data by branch if a specific branch is selected
        # Before filtering
        print(f"User Data List (Before): {user_data_list}")

        # Filter user data by branch if a specific branch is selected
        filtered_user_data_list = [user_data for user_data in user_data_list if user_data['Branch'] == selected_branch]

        # After filtering
        print(f"User Data List (After): {filtered_user_data_list}")

    return render_template('admin_portal.html', user_data_list=filtered_user_data_list,
                           user_credentials_list=user_credentials_list)


@app.route('/download_company_data/<company>/<company_type>', methods=['GET'])
def download_company_data(company, company_type):
    directory = f"{company_type}_Companies"
    file_path = os.path.join(directory, f"{company}.xlsx")

    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "Company data not found", 404

@app.route('/serve_company_data/<company_type>/<company_name>', methods=['GET'])
def serve_company_data(company_type, company_name):
    directory = f"{company_type}_Companies"
    file_path = os.path.join(directory, f"{company_name}.xlsx")

    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "Company data not found", 404

# Route for serving static files (Bootstrap CSS)
@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)


@app.route('/data_download.html', methods=['GET'])
def data_download():
    return render_template('data_download.html')

def fetch_user_data(file_name):
    user_data_list = []
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            # Check if the row has valid data (not placeholders)
            if row and row[0] is not None and row[0] != 'Roll Number':
                user_data = {
                    "Roll Number": row[0],
                    "Name": row[1],
                    "Branch": row[2],
                    "Phone Number": row[3],
                    "Email ID": row[4],
                    "Official Email ID": row[5]
                }
                user_data_list.append(user_data)
    return user_data_list



def fetch_user_credentials(file_name):
    user_credentials_list = []
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row:
                user_credential = {
                    "Username": row[0],
                    "Password": row[1]
                }
                user_credentials_list.append(user_credential)
    return user_credentials_list


@app.route('/admin.html')
def admin():
    return render_template('admin.html')



def is_roll_number_registered(roll_number, file_name):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row and row[0] == roll_number:
                return True
    return False

def save_user_details(details, file_name):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Roll Number', 'Name', 'Branch', 'Phone Number', 'Email ID', 'Official Email ID'])

    sheet.append([details['roll_number'], details['name'], details['branch'],
                  details['phone_number'], details['email_id'], details['official_email_id']])
    workbook.save(file_name)

def save_credentials(username, password, file_name):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Username', 'Password'])

    # Check if the username (roll number) already exists in the sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row and row[0] == username:
            workbook.save(file_name)  # Save the workbook
            return "This roll number has already been registered"

    # If the username doesn't exist, add it
    sheet.append([username, password])
    workbook.save(file_name)  # Save the workbook
    return "Registration successful"


def validate_credentials(username, password):
    credentials_file = 'user_credentials.xlsx'
    if not os.path.exists(credentials_file):
        return False

    workbook = load_workbook(credentials_file)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        stored_username, stored_password = row[0], row[1]
        if stored_username is not None and stored_password is not None:
            # Convert both username and password to strings before stripping
            stored_username = str(stored_username).strip().lower()
            stored_password = str(stored_password)

            # Compare username case-insensitively and password as-is
            if stored_username == username.strip().lower() and stored_password == password:
                return True

    return False


@app.route('/login', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']

    if validate_credentials(username, password):
        # Redirect to company_selection with the roll_number parameter
        return redirect(url_for('company_selection', roll_number=username))
    else:
        return "Invalid login credentials", 401


@app.route('/company_selection/<roll_number>', methods=['GET', 'POST'])
def company_selection(roll_number):
    if request.method == 'POST':
        company_type = request.form.get('company_type')
        # Use either 'it_companies' or 'non_it_companies' based on the selected company type
        companies = it_companies if company_type == 'IT' else non_it_companies
        return render_template('companies.html', roll_number=roll_number, company_type=company_type, companies=companies)
    return render_template('company_selection.html', roll_number=roll_number)

@app.route('/companies/<roll_number>/<company_type>', methods=['GET', 'POST'])
def companies(roll_number, company_type):
    if request.method == 'POST':
        company_name = request.form['company_name']
        status = update_excel(roll_number, company_type, company_name)
        return status
    return render_template('companies.html', roll_number=roll_number, company_type=company_type)

def update_excel(roll_number, company_type, company_name):
    directory = f"{company_type}_Companies"
    if not os.path.exists(directory):
        os.makedirs(directory)

    file_path = os.path.join(directory, f"{company_name}.xlsx")

    if not os.path.exists(file_path):
        # Create a new workbook and sheet if the Excel file does not exist
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Roll Number'])
    else:
        # Load the existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

    # Check if the roll number already exists in the sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row and row[0] == roll_number:
            return "exists"

    # If the roll number doesn't exist, append it
    sheet.append([roll_number])

    # Save the workbook
    workbook.save(file_path)
    return "added"

@app.route('/get_user_details/<roll_number>', methods=['GET'])
def get_user_details(roll_number):
    user_details = fetch_user_details(roll_number, 'user_details.xlsx')
    return jsonify(user_details)

def fetch_user_details(roll_number, file_name):
    user_details = {}
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row and row[0] == roll_number:
                user_details = {
                    "Name": row[1],
                    "Branch": row[2],
                    "Phone Number": row[3],
                    "Email ID": row[4],
                    "Official Email ID": row[5]
                }
                break
    return user_details



import os

@app.route('/display_directory_contents/<directory_type>', methods=['GET'])
def display_directory_contents(directory_type):
    directory = f"{directory_type}"
    file_list = []

    if os.path.exists(directory):
        file_list = os.listdir(directory)

    return render_template('data_download.html', directory_contents=file_list)


if __name__ == "__main__":
    app.run(debug=True)
